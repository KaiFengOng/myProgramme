import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

wb = Workbook()
wb = openpyxl.load_workbook("StudentFinalScore.xlsx")
res = len(wb.sheetnames)

# to show all the sheet that created in excel file and print out the sheet name
sheets = wb.sheetnames
print(sheets)
print("Your current total sheet is " + str(res))

# this part only allow teacher/lec to create a list of sheet for them to enter student score
if res < 10:
    print("Please Note that you can only create a 10 total amount sheet of File")

    totalSheet = input("Enter total Num of sheet you want to create:")
    if totalSheet != '0' and totalSheet > '10':
        for i in range(int(totalSheet)):
            sheetName = input("Enter the sheet Name for " + str(i) + ":")
            ws1 = wb.create_sheet(sheetName)
            ws1.title = sheetName
            totalSheet += '1'

    else:
        print("Please enter below 10 to create your excel sheet")

else:
    # before prompt user to enter sheet name, cal the existing file if max or no
    # if max then need aware uer that sheet reach max
    print("Oops, you have reach the maximum of the sheet that you can create")

wb.save('StudentFinalScore.xlsx')