from openpyxl import Workbook, load_workbook

# Enter student score and store in data in excel file
import pandas as pd

wb = load_workbook('StudentFinalScore.xlsx')
# wb = pd.read_excel('StudentFinalScore.xlsx')
# data = pd.read_excel()

# to show all the sheet that created in excel file
sheets = wb.sheetnames
print(sheets)

choice = input("Enter the correct sheet name you want to access:")

# Check if the chosen sheet name is valid**
if choice in sheets:
    # Get the selected sheet
    active_sheet = wb[choice]

    totalStudent = int(input("Please enter total amount of your student:"))

    # check if the number of student is valid
    if totalStudent > 0 and (choice == 'Math' and totalStudent <= 31) or (choice != 'Math' and totalStudent <= 30):
        for i in range(totalStudent):
            studentName = input("Enter Student Name:")
            studentScore = input("Enter Student Score:")

            # Write column headers if they don't exist
            if active_sheet['A1'].value is None:
                active_sheet['A1'].value = "Student Name"
                active_sheet['B1'].value = "Score"

            # Append student data to the sheet
            active_sheet.append([studentName, studentScore])
        wb.save('StudentFinalScore.xlsx')

    else:
        print("Please Enter the correct amount of student")

else:
    print("Please select the correct sheet name")