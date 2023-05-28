import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

wb = Workbook()
wb = openpyxl.load_workbook("StudentFinalScore.xlsx")

# to show all the sheet that created in excel file
sheets = wb.sheetnames
print(sheets)

# to get the active sheet
print(wb.active.title)

wb.save('StudentFinalScore.xlsx')